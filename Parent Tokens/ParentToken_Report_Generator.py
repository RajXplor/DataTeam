"""
DEVELOPED BY 7GONEINSANE
PARENT TOKEN & NO BANKING REPORT GENERATOR
===========================================
INPUT FILES (auto-detected by name — place in same folder):
  - *payment_plan*        e.g. Larmenier_OSHC_payment_plan_import.csv
  - *DS*TOKEN* / *token*  e.g. Larmenier_OSHC_DS_TOKENS.csv
  - *guardian_financial*  e.g. guardian_financial_account_list_...csv
  - *parent_bank*         e.g. parent_bank_details_summary_report_...csv (optional)

OUTPUTS:
  - ParentToken_Import_{ServiceName}.csv
  - {ServiceName}_DuplicateGateway_Review.xlsx
  - {ServiceName}_No_BankingReport.xlsx
"""

import sys
import re
from pathlib import Path

if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

W = 66

def divider():        print("  " + "─" * (W + 2))
def step(n, t, lbl):  print(f"\n  🔹 [{n}/{t}]  {lbl}")
def info(lbl, val=""):
    print(f"       {lbl} {'·'*max(2,46-len(lbl))} {val}" if val else f"       {lbl}")
def warn(msg):  print(f"       ⚠️   {msg}")
def good(msg):  print(f"       ✅  {msg}")
def err(msg):   print(f"       ❌  {msg}")

print()
print("  " + "━" * (W + 2))
print("  💳  PARENT TOKENS  &  NO BANKING REPORT GENERATOR")
print("  " + "━" * (W + 2))

# ─────────────────────────────────────────────────────────────
# STEP 1 — PACKAGES
# ─────────────────────────────────────────────────────────────
step(1, 6, "Checking packages")
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    good("pandas + openpyxl are ready")
except ImportError as e:
    err(f"Missing package: {e}")
    err("Fix: pip install pandas openpyxl --break-system-packages")
    print()
    input("  Press Enter to exit...")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────
# STEP 2 — AUTO-DETECT INPUT FILES
# ─────────────────────────────────────────────────────────────
step(2, 6, "Detecting input files 📂")

script_folder = Path(__file__).parent


def find_file(folder, *keywords):
    """Find first file whose lowercase name contains ALL keywords."""
    valid_ext = {".csv", ".xlsx", ".xlsm", ".xls"}
    for f in sorted(folder.iterdir()):
        if f.suffix.lower() not in valid_ext:
            continue
        name_lower = f.name.lower()
        if all(k.lower() in name_lower for k in keywords):
            return f
    return None


def read_any(path):
    """
    Read CSV or Excel into DataFrame.
    Handles CSV files with a 1–2 line report-title preamble automatically.
    """
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        return pd.read_excel(path, dtype=str)
    with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
        first_line = fh.readline()
    has_preamble = (
        first_line.startswith('"')
        and "," not in first_line.replace('"', "").strip()
    )
    skip = 2 if has_preamble else 0
    return pd.read_csv(path, dtype=str, skiprows=skip, encoding="utf-8-sig")


pp_path  = find_file(script_folder, "payment_plan")
ds_path  = (find_file(script_folder, "ds", "token")
            or find_file(script_folder, "ds_token"))
gfl_path = find_file(script_folder, "guardian_financial")

missing = False
for label, path in [
    ("Payment plan file  ", pp_path),
    ("DS Tokens file     ", ds_path),
    ("Guardian list file ", gfl_path),
]:
    if path:
        good(f"{label} → {path.name}")
    else:
        err(f"{label} → NOT FOUND")
        missing = True

if missing:
    print()
    warn(f"Place missing files in: {script_folder}")
    print()
    input("  Press Enter to exit...")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────
# STEP 3 — LOAD DATA
# ─────────────────────────────────────────────────────────────
step(3, 6, "Loading data 📊")

pp  = read_any(pp_path);  pp.columns  = pp.columns.str.strip()
ds  = read_any(ds_path);  ds.columns  = ds.columns.str.strip()
gfl = read_any(gfl_path); gfl.columns = gfl.columns.str.strip()

info("Payment plan rows ", f"{len(pp):,}")
info("DS token rows     ", f"{len(ds):,}")
info("Guardian list rows", f"{len(gfl):,}")

service_name = (pp["Service_Name"].dropna().iloc[0].strip()
                if "Service_Name" in pp.columns else "Service")
service_id   = (pp["Service_ID"].dropna().iloc[0].strip()
                if "Service_ID"  in pp.columns else "")
info("Service name      ", service_name)

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def cv(val):
    """Clean value — return stripped string or '' for NaN/None."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def norm(s):
    """Lowercase + collapse whitespace for fuzzy matching."""
    return re.sub(r"\s+", " ", cv(s).lower()).strip()


def child_in_gfl(pp_child, gfl_children_csv):
    """True when pp_child appears in the comma-separated GFL children list."""
    return norm(pp_child) in [norm(c) for c in gfl_children_csv.split(",")]


def ds_client_to_norm(client_str):
    """DS Client is stored as 'Last, First' — convert to 'first last'."""
    s = cv(client_str)
    if "," in s:
        last, first = s.split(",", 1)
        return norm(first.strip() + " " + last.strip())
    return norm(s)

# ─────────────────────────────────────────────────────────────
# BUILD LOOKUP TABLES
# ─────────────────────────────────────────────────────────────
gfl["_hn"] = gfl["Account Holder"].apply(norm)
gfl_by_name: dict = {}
for _, row in gfl.iterrows():
    gfl_by_name.setdefault(row["_hn"], []).append(row)

ds["_cu"] = ds["Club Number"].apply(lambda v: cv(v).upper())
ds_by_gw: dict = {}
for _, row in ds.iterrows():
    key = row["_cu"]
    if key and not key.startswith("XXX"):
        ds_by_gw[key] = row

ds_by_client: dict = {}
for _, row in ds.iterrows():
    key = ds_client_to_norm(cv(row["Client"]))
    if key:
        ds_by_client.setdefault(key, []).append(row)

pp["_pf"] = pp["Parent_First_Name"].apply(cv) + " " + pp["Parent_Last_Name"].apply(cv)
pp["_cf"] = pp["Child_First_Name"].apply(cv)  + " " + pp["Child_Last_Name"].apply(cv)
pp["_pn"] = pp["_pf"].apply(norm)
pp["_cn"] = pp["_cf"].apply(norm)
pp["_gu"] = pp["Gateway_Reference"].apply(lambda v: cv(v).upper())

# ─────────────────────────────────────────────────────────────
# STEP 4 — PROCESS PAYMENT PLAN
# ─────────────────────────────────────────────────────────────
step(4, 6, "Processing payment plan rows 🔄")

valid_tokens     = []
iss_no_gw        = []
iss_no_parent    = []
iss_child_diff   = []
dup_gateway_rows = []

gw_per_parent  = pp.groupby("_pn")["_gu"].nunique()
dup_parent_set = set(gw_per_parent[gw_per_parent > 1].index)

for i, row in pp.iterrows():
    csv_row     = i + 2
    parent_full = cv(row["_pf"])
    parent_norm = row["_pn"]
    child_full  = cv(row["_cf"])
    gateway     = row["_gu"]
    svc_display = cv(row.get("Service_Name", service_name))

    if parent_norm in dup_parent_set:
        gfl_id = (cv(gfl_by_name[parent_norm][0]["ID"])
                  if parent_norm in gfl_by_name else "NOT IN GFL")
        dup_gateway_rows.append({
            "Service Name"    : svc_display,
            "Xplor ParentID"  : gfl_id,
            "Parent Full Name": parent_full,
            "Child Name"      : child_full,
            "Gateway Reference": cv(row["Gateway_Reference"]),
            "Review Note"     : "DUPLICATE — same parent has more than one gateway reference",
        })
        continue

    if gateway not in ds_by_gw:
        iss_no_gw.append({
            "PP Row": csv_row, "Parent Full Name": parent_full,
            "Child Name": child_full, "Gateway Ref": cv(row["Gateway_Reference"]),
            "Service": svc_display,
            "Issue": "Gateway reference not found in DS Tokens",
        })
        continue

    if parent_norm not in gfl_by_name:
        child_hits = [r for _, r in gfl.iterrows()
                      if row["_cn"] in norm(cv(r["Child Names"]))]
        if child_hits:
            reason = (f"Child '{child_full}' found under "
                      f"'{cv(child_hits[0]['Account Holder'])}' "
                      f"(ID {cv(child_hits[0]['ID'])}) — parent name differs")
        else:
            reason = (f"Parent '{parent_full}' and child '{child_full}' "
                      f"not found in guardian list at all")
        dsr = ds_by_gw[gateway]
        iss_no_parent.append({
            "PP Row": csv_row, "Parent Full Name": parent_full,
            "Child Name": child_full, "Gateway Ref": cv(row["Gateway_Reference"]),
            "DS Token": cv(dsr["Adfit No"]), "Service": svc_display, "Issue": reason,
        })
        continue

    gfl_match = gfl_by_name[parent_norm][0]
    gfl_kids  = cv(gfl_match["Child Names"])
    gfl_pid   = cv(gfl_match["ID"])

    if not child_in_gfl(child_full, gfl_kids):
        iss_child_diff.append({
            "PP Row": csv_row, "Parent Full Name": parent_full,
            "Child in PP": child_full, "GFL Children": gfl_kids,
            "GFL Parent ID": gfl_pid, "Gateway Ref": cv(row["Gateway_Reference"]),
            "Service": svc_display,
            "Issue": "Parent matched but child name does not match GFL",
        })
        continue

    dsr = ds_by_gw[gateway]
    valid_tokens.append({"Parent ID": gfl_pid, "Token": cv(dsr["Adfit No"])})

# ── Console summary ───────────────────────────────────────────
print()
info("✅ Valid token imports          ", f"{len(valid_tokens):,}")
info("🔁 Duplicate gateways (review) ", f"{len(dup_gateway_rows):,}")
info("❌ Gateways not in DS           ", f"{len(iss_no_gw):,}")
info("❌ Parents not in GFL           ", f"{len(iss_no_parent):,}")
info("❌ Child mismatches             ", f"{len(iss_child_diff):,}")

if iss_no_gw:
    print(); divider(); info(" 🔍 GATEWAYS NOT FOUND IN DS TOKENS"); divider()
    for x in iss_no_gw:
        print(f"       Row {x['PP Row']:>3}  {x['Parent Full Name']:<28}  {x['Child Name']}")
        print(f"              Gateway: {x['Gateway Ref']}")

if iss_no_parent:
    print(); divider(); info(" 🔍 PARENTS / CHILDREN NOT IN GUARDIAN LIST"); divider()
    for x in iss_no_parent:
        print(f"       Row {x['PP Row']:>3}  {x['Parent Full Name']:<28}  {x['Child Name']}")
        print(f"              {x['Issue']}")

if iss_child_diff:
    print(); divider(); info(" 🔍 CHILD NAME MISMATCHES"); divider()
    for x in iss_child_diff:
        print(f"       Row {x['PP Row']:>3}  {x['Parent Full Name']:<28}  PP: {x['Child in PP']}")
        print(f"              GFL children: {x['GFL Children']}")

if iss_no_gw or iss_no_parent or iss_child_diff:
    divider()

# ─────────────────────────────────────────────────────────────
# STEP 5 — WRITE OUTPUT FILES
# ─────────────────────────────────────────────────────────────
step(5, 6, "Writing output files 💾")

safe_svc = service_name.replace(" ", "_").replace("/", "-")

FONT_NAME = "Aptos"
FONT_SIZE = 11

def mk_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header(cell, bg="1F3864", fg="FFFFFF"):
    """Dark navy header — Aptos 11 bold white, centred."""
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name=FONT_NAME, bold=True, color=fg, size=FONT_SIZE)
    cell.border    = mk_border()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_data(cell, bg="FFFFFF", bold=False):
    """Data cell — Aptos 11, centred, with light border."""
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name=FONT_NAME, bold=bold, size=FONT_SIZE)
    cell.border    = mk_border()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws):
    col_max: dict = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col_letter = get_column_letter(cell.column)
            text_len   = len(str(cell.value))
            if text_len > col_max.get(col_letter, 0):
                col_max[col_letter] = text_len

    for col_letter, max_len in col_max.items():
        width = max(10, int(max_len * 1.1) + 4)
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows():
        row_num   = row[0].row
        max_lines = 1
        for cell in row:
            if cell.value:
                col_letter     = get_column_letter(cell.column)
                col_w          = max(1, ws.column_dimensions[col_letter].width)
                text           = str(cell.value)
                chars_per_line = max(1, int(col_w / 1.1))
                lines = max(1, -(-len(text) // chars_per_line))
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row_num].height = max(20, max_lines * 15 + 6)

# ── A) ParentToken_Import_{ServiceName}.csv ───────────────────
token_out = script_folder / f"ParentToken_Import_{safe_svc}.csv"
if valid_tokens:
    pd.DataFrame(valid_tokens).to_csv(token_out, index=False, encoding="utf-8-sig")
    good(f"ParentToken_Import_{safe_svc}.csv  ({len(valid_tokens)} rows)")
else:
    warn("No valid rows — ParentToken_Import CSV was not created")

# ── B) Duplicate Gateway Review (xlsx) ───────────────────────
if dup_gateway_rows:
    dup_out  = script_folder / f"{safe_svc}_DuplicateGateway_Review.xlsx"
    wb_dup   = Workbook()
    ws_dup   = wb_dup.active
    ws_dup.title = "Duplicate Gateway Review"

    dup_display_cols = [
        "Service Name", "Xplor ParentID", "Parent Full Name",
        "Child Name", "Gateway Reference", "Review Note",
    ]

    for ci, h in enumerate(dup_display_cols, 1):
        apply_header(ws_dup.cell(row=1, column=ci, value=h))

    for ri, rd in enumerate(dup_gateway_rows, 2):
        row_bg = "FFF9E6" if ri % 2 == 0 else "FFFEF7"
        for ci, col in enumerate(dup_display_cols, 1):
            val = rd.get(col, "")
            c   = ws_dup.cell(row=ri, column=ci, value=val)
            if col == "Review Note":
                apply_data(c, bg="FFCC00", bold=True)
            else:
                apply_data(c, bg=row_bg)

    autofit(ws_dup)
    ws_dup.freeze_panes = "A2"
    wb_dup.save(dup_out)
    good(f"{dup_out.name}  ({len(dup_gateway_rows)} rows)")

# ─────────────────────────────────────────────────────────────
# STEP 6 — NO BANKING REPORT
# ─────────────────────────────────────────────────────────────
step(6, 6, "No Banking Report 🏦")
print()
divider()
print("  📁 Is the parent_bank_details_summary_report")
print("     uploaded to this folder?")
divider()
print()
response = input("  Enter Yes or No: ").strip().lower()
print()

if response not in ("yes", "y"):
    info("Banking report skipped 🚫")
else:
    bank_path = (find_file(script_folder, "parent_bank")
                 or find_file(script_folder, "bank_detail"))

    if not bank_path:
        err("parent_bank_details_summary_report file not found — report skipped")
        err(f"Expected in: {script_folder}")
    else:
        good(f"Bank file → {bank_path.name}")
        bank = read_any(bank_path)
        bank.columns = bank.columns.str.strip()
        info("Bank rows (total)     ", f"{len(bank):,}")

        bank_detail_col = None
        for col in bank.columns:
            if "bank detail" in col.lower():
                bank_detail_col = col
                break

        if bank_detail_col is None:
            warn("Could not find 'Bank Details (Y/N)' column — using all rows")
            no_bank_df = bank.copy()
        else:
            no_bank_df = bank[bank[bank_detail_col].str.strip().str.lower() == "no"].copy()

        no_bank_df = no_bank_df.reset_index(drop=True)
        info("Parents without banking", f"{len(no_bank_df):,}")

        def get_note(parent_full_name):
            """Return (note_text, hex_colour) for this parent."""
            pn = norm(parent_full_name)
            ds_ents = ds_by_client.get(pn, [])

            if not ds_ents:
                return ("Not found", "")

            def is_cancelled(dr):
                adfit = cv(dr.get("Adfit No", "")).lower()
                club  = cv(dr.get("Club Number", "")).upper()
                return adfit in ("n/a", "na", "") or club.startswith("XXX")

            all_cancelled = all(is_cancelled(dr) for dr in ds_ents)
            has_valid     = any(not is_cancelled(dr) for dr in ds_ents)

            if all_cancelled:
                return ("Cancelled - No billing since 31/12/2019", "F4B942")  # orange
            if has_valid:
                return ("Review", "FFFF00")                                    # yellow
            return ("Not found", "")

        no_bank_rows = []
        for _, br in no_bank_df.iterrows():

            first   = cv(br.get("First Name", ""))
            last    = cv(br.get("Last Name", ""))
            pfull   = f"{first} {last}".strip()
            svc_id  = cv(br.get("Service ID", service_id))
            svc_nm  = cv(br.get("Service Name", service_name))

            pn        = norm(pfull)
            gfl_match = gfl_by_name.get(pn, [])
            children  = cv(gfl_match[0]["Child Names"]) if gfl_match else ""

            note_text, note_color = get_note(pfull)

            no_bank_rows.append({
                "Service ID"       : svc_id,
                "Service Name"     : svc_nm,
                "Parent Full Name" : pfull,
                "Children Full Name": children,
                "Notes"            : note_text,
                "_note_color"      : note_color,
                "Gateway Reference": "",
            })

        nb_out = script_folder / f"{safe_svc}_No_BankingReport.xlsx"
        wb_nb  = Workbook()
        ws_nb  = wb_nb.active
        ws_nb.title = "No Banking Report"

        nb_display_cols = [
            "Service ID", "Service Name", "Parent Full Name",
            "Children Full Name", "Notes", "Gateway Reference",
        ]

        for ci, h in enumerate(nb_display_cols, 1):
            apply_header(ws_nb.cell(row=1, column=ci, value=h))

        for ri, rd in enumerate(no_bank_rows, 2):
            row_bg = "EEF3FB" if ri % 2 == 0 else "FFFFFF"
            for ci, col in enumerate(nb_display_cols, 1):
                val = rd.get(col, "")
                c   = ws_nb.cell(row=ri, column=ci, value=val)
                if col == "Notes":
                    nc = rd.get("_note_color", "")
                    apply_data(c, bg=nc if nc else row_bg, bold=bool(nc))
                else:
                    apply_data(c, bg=row_bg)

        autofit(ws_nb)
        ws_nb.freeze_panes = "A2"
        wb_nb.save(nb_out)
        good(f"{nb_out.name}  ({len(no_bank_rows)} rows)")

# ─────────────────────────────────────────────────────────────
# FINAL SUMMARY
# ─────────────────────────────────────────────────────────────
print()
print("  " + "━" * (W + 2))
print("  ✅  SUCCESS  —  Script finished without errors")
print("  🚀  DEVELOPED BY 7GONEINSANE")
print("  " + "━" * (W + 2))
print()
print(f"  📄 Results:")
print(f"       ✅  Valid tokens in ParentToken_Import_{safe_svc}.csv  :  {len(valid_tokens)}")
print(f"       🔁  Duplicate gateway rows (review file)              :  {len(dup_gateway_rows)}")
print()
print(f"  ❌ Excluded from ParentToken_Import:")
print(f"       Gateway not found in DS Tokens     :  {len(iss_no_gw)}")
print(f"       Parent not found in guardian list  :  {len(iss_no_parent)}")
print(f"       Child name does not match          :  {len(iss_child_diff)}")
print()
input("  Press Enter to close...")